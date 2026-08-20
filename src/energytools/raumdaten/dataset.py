"""energytools.raumdaten.dataset -- loading, store and extraction (API reference part 03, section 2).

Loading (`load_dataset`, `DatasetStore`): JSON package + JSON Schema -> frozen
:class:`~energytools.raumdaten.model.Dataset`.  Extraction (`DatasetExtractor`):
stage-0 pipeline that reads a **copy** of the source workbook deterministically
(cell graph: values + formulas + cached results), checksums it, validates the
result against the package JSON Schema and writes the canonical JSON package.
No cell is written, no macro runs, no link is followed.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from pathlib import Path
from typing import Any

from energytools.common.errors import (
    DatasetNotFoundError,
    DatasetValidationError,
    UnitError,
)
from energytools.common.language import TrilingualText
from energytools.common.provenance import Provenance, SourceRef
from energytools.common.units import Quantity, Unit
from energytools.common.versioning import ChangelogEntry, DatasetRelease
from energytools.raumdaten.model import (
    AreaTable,
    BuildingCategoryMapping,
    CategoryTable,
    ClimateData,
    ClimateStation,
    Dataset,
    DesignDaySeries,
    FullLoadHoursTable,
    HourlyProfile,
    MonthlyProfile,
    Parameter,
    ParameterValue,
    QhcTable,
    RoomUse,
    RoomUseInputs,
    RoomUseProfile,
    RoomUseSchedule,
    Sia2028Monthly,
    Sia3801Coefficients,
    Sia3801Result,
    ValueKind,
    normalize_room_use_code,
)
from energytools.raumdaten.schema import PACKAGE_SCHEMA

__all__ = ["DEFAULT_DATASET_DIR_ENV", "DatasetExtractor", "DatasetStore", "load_dataset"]

DEFAULT_DATASET_DIR_ENV = "ENERGYTOOLS_DATASET_DIR"
_RELEASE_ID_RE = re.compile(r"^[A-Za-z0-9._-]+$")

# Process-wide cache: (release_id, dataset_dir) -> Dataset.  Enforces the
# invariant "one frozen Dataset per release id" per dataset directory.
_LOAD_CACHE: dict[tuple[str, str], Dataset] = {}


def default_dataset_dir() -> str:
    """The configured dataset directory (``ENERGYTOOLS_DATASET_DIR`` or ``./data/datasets``)."""
    return os.environ.get(DEFAULT_DATASET_DIR_ENV, "data/datasets")


def _canonical_bytes(data: dict) -> bytes:
    """Deterministic canonical JSON bytes (checksum basis)."""
    return json.dumps(
        data, sort_keys=True, ensure_ascii=False, separators=(",", ":"), allow_nan=False
    ).encode("utf-8")


def _content_checksum(data: dict) -> str:
    """SHA-256 of the canonical package content, excluding the checksum field itself."""
    without_checksum = json.loads(json.dumps(data))
    release = without_checksum.get("release")
    if isinstance(release, dict):
        release.pop("checksum_sha256", None)
    return hashlib.sha256(_canonical_bytes(without_checksum)).hexdigest()


def _schema_errors(instance: Any) -> list[str]:
    """Human-readable JSON-Schema validation errors (empty when valid)."""
    # Lazily imported: jsonschema is part of the 'data' extra.
    import jsonschema  # type: ignore[import-untyped]

    validator = jsonschema.Draft202012Validator(PACKAGE_SCHEMA)
    return sorted(
        f"{list(error.path) or ['<root>']}: {error.message}"
        for error in validator.iter_errors(instance)
    )


def _require_jsonschema() -> None:
    try:
        import jsonschema  # noqa: F401
    except ImportError:
        raise ImportError(
            "jsonschema is required to load dataset packages; "
            "install it with `pip install 'energytools[data]'`"
        ) from None


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def load_dataset(release_id: str, path: str | None = None) -> Dataset:
    """Load a dataset release from disk (JSON package + JSON Schema).

    Results are cached in the process-wide :class:`DatasetStore`; subsequent
    calls with the same ``release_id`` (and dataset directory) return the
    identical frozen object.

    Args:
        release_id: Release id (e.g. ``"V221"``).
        path: Directory of the package; defaults to the configured dataset
            directory (``ENERGYTOOLS_DATASET_DIR`` or ``./data/datasets``).

    Raises:
        DatasetNotFoundError: release not installed.
        DatasetValidationError: package fails schema/checksum/content validation
            (a corrupt/foreign package is never half-loaded).
    """
    dataset_dir = path or default_dataset_dir()
    return DatasetStore(dataset_dir).get(release_id)


class DatasetStore:
    """Process-wide registry of loaded releases.

    Enforces the invariant "one frozen :class:`Dataset` per release id" and
    answers existence queries without touching disk.

    Args:
        dataset_dir: Root of installed packages (defaults to the configured
            dataset directory).
    """

    def __init__(self, dataset_dir: str | None = None) -> None:
        self.dataset_dir = dataset_dir or default_dataset_dir()

    # -- public API ---------------------------------------------------------

    def get(self, release_id: str) -> Dataset:
        """Load-on-demand (via :func:`load_dataset` semantics), cached.

        Raises:
            DatasetNotFoundError: release not installed.
            DatasetValidationError: package fails validation.
        """
        self._check_release_id(release_id)
        key = (release_id, os.path.abspath(self.dataset_dir))
        cached = _LOAD_CACHE.get(key)
        if cached is not None:
            return cached
        dataset = self._load_from_disk(release_id)
        _LOAD_CACHE[key] = dataset
        return dataset

    def list(self) -> list[DatasetRelease]:
        """Installed releases (from package manifests), newest first."""
        releases = []
        root = Path(self.dataset_dir)
        if not root.is_dir():
            return []
        for package_file in sorted(root.glob("*/package.json")):
            try:
                data = json.loads(package_file.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue  # a corrupt package is skipped here; get() reports it properly
            release = data.get("release")
            if not isinstance(release, dict) or not release.get("id"):
                continue
            try:
                releases.append(_release_from_dict(release))
            except (KeyError, TypeError, ValueError):
                continue
        return sorted(releases, key=lambda rel: (rel.publication_date, rel.id), reverse=True)

    def register(self, dataset: Dataset) -> None:
        """Pre-register an in-memory dataset (tests, custom packages).

        Raises:
            ValueError: on a duplicate id with different content.
        """
        key = (dataset.release_id, os.path.abspath(self.dataset_dir))
        existing = _LOAD_CACHE.get(key)
        if existing is not None and existing != dataset:
            raise ValueError(
                f"dataset release '{dataset.release_id}' is already registered with different content"
            )
        _LOAD_CACHE[key] = dataset

    def refresh(self) -> None:
        """Drop the cache and re-scan the dataset directory."""
        prefix = os.path.abspath(self.dataset_dir)
        for key in [k for k in _LOAD_CACHE if k[1] == prefix]:
            del _LOAD_CACHE[key]

    # -- internals ----------------------------------------------------------

    @staticmethod
    def _check_release_id(release_id: str) -> None:
        if not release_id or not _RELEASE_ID_RE.match(release_id):
            raise DatasetValidationError(
                f"invalid release id '{release_id}' (allowed: letters, digits, '.', '_', '-')",
                {"release_id": release_id},
            )

    def _load_from_disk(self, release_id: str) -> Dataset:
        release_dir = Path(self.dataset_dir) / release_id
        package_file = release_dir / "package.json"
        if not package_file.is_file():
            raise DatasetNotFoundError(release_id)

        _require_jsonschema()

        try:
            data = json.loads(package_file.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise DatasetValidationError(
                f"package '{release_id}' is not valid JSON: {exc}",
                {"release_id": release_id, "errors": [str(exc)]},
            ) from exc
        except OSError as exc:
            raise DatasetValidationError(
                f"package '{release_id}' cannot be read: {exc}",
                {"release_id": release_id, "errors": [str(exc)]},
            ) from exc

        errors = _schema_errors(data)
        if errors:
            raise DatasetValidationError(
                f"package '{release_id}' fails schema validation",
                {"release_id": release_id, "errors": errors},
            )

        expected_checksum = data.get("release", {}).get("checksum_sha256")
        actual_checksum = _content_checksum(data)
        if expected_checksum and expected_checksum.lower() != actual_checksum.lower():
            raise DatasetValidationError(
                f"package '{release_id}' checksum mismatch "
                f"(expected {expected_checksum}, computed {actual_checksum})",
                {"release_id": release_id, "errors": ["checksum mismatch"]},
            )

        try:
            return Dataset.from_package_dict(data)
        except (ValueError, TypeError, KeyError) as exc:
            raise DatasetValidationError(
                f"package '{release_id}' has inconsistent content: {exc}",
                {"release_id": release_id, "errors": [str(exc)]},
            ) from exc


def _release_from_dict(data: dict) -> DatasetRelease:
    from datetime import date

    return DatasetRelease(
        id=data["id"],
        edition=data["edition"],
        publication_date=date.fromisoformat(data["publication_date"]),
        checksum_sha256=data["checksum_sha256"],
        source_workbook=data["source_workbook"],
        extraction_tool_version=data["extraction_tool_version"],
        supersedes=data.get("supersedes"),
        changelog=tuple(
            ChangelogEntry(
                version=entry["version"],
                date=date.fromisoformat(entry["date"]),
                change=entry["change"],
                migration=entry.get("migration"),
            )
            for entry in data.get("changelog", [])
        ),
    )


# ---------------------------------------------------------------------------
# Extraction (stage 0)
# ---------------------------------------------------------------------------

# Standard versions of the Volll_Lüft table (assessment 1.2).  V221 ships one
# version block; the version axis is the model for the dataset's evolution.
STANDARD_VERSIONS = ("SIA 2024:2015", "prSIA 2024:2021", "prSIA 2024-C1:2024")
REQUIRED_SHEETS = (
    "Eingabedaten",
    "Begriffe",
    "Datenblatt",
    "Profile",
    "Monatswerte",
    "Winter_Auslegung",
    "Aug_Auslegung",
    "Volll_Lüft",
    "Qhc_Klimastat",
    "Fläche-E",
    "GEPAMOD",
    "SIA 380-1",
)

_SECTION_HEADERS = (
    "Raum",
    "Bauphysikalische Eigenschaften",
    "Raumklima",
    "Schallschutz",
    "Personen",
    "Geräte und Prozessanlagen",
    "Beleuchtung",
    "Lüftung",
    "Raumkühlung",
    "Befeuchtung",
    "Raumheizung",
    "Wasser",
)

_ERROR_VALUE_STRINGS = {"#N/A", "#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "Fehler"}

# Monatswerte: 11-row blocks per station, starting at rows 1, 12, 23, ...
_MONTHLY_BLOCK_UNITS = {
    "t_aussen": "°C",
    "radiation_horizontal": "MJ/m2",
    "radiation_east": "MJ/m2",
    "radiation_south": "MJ/m2",
    "radiation_west": "MJ/m2",
    "radiation_north": "MJ/m2",
    "precipitation": "mm",
    "mixing_ratio": "g/kg",
    "absolute_humidity": "g/m3",
}

# ---------------------------------------------------------------------------
# Per-category reference tables (batch C)
# ---------------------------------------------------------------------------

#: Per-category blocks of the ``Fläche-E`` family (rows 3-47, one metric row
#: each, block boundaries follow the row-1 title cells).  Columns are
#: 1-based; ``kind`` is the produced :class:`CategoryTable` kind (``None`` for
#: the AE block whose kind depends on the sheet variant).  Blocks whose rows
#: continue with a redundant sub-block stop at the first section header
#: (``stop_at_section_header``): the AE block's "SIA 2024 Zielwerte/Bestand"
#: sub-sections duplicate the Fläche-ZW / Fläche-Best sheets, and the DB
#: block's "SIA 390:2021" sub-section is a different standard.
_FLAECHE_BLOCKS = (
    # label, unit, first category column, last category column, kind, stop_at_header
    (29, 30, 31, 52, None, True),  # AC/AD, AE..AZ — SIA 2024 Standardwerte
    (55, 56, 57, 78, "sia3801_tab27", False),  # BC/BD, BE..BZ — SIA 380/1 Tab. 27
    (80, 81, 82, 103, "harmonized", False),  # CB/CC, CD..CY
    (106, 107, 108, 127, "sia2040", True),  # DB/DC, DD..DW
    (130, 131, 132, 152, "weighted_energy", False),  # DZ/EA, EB..EV — Minergie 2017
    (154, 155, 156, 177, "minergie", False),  # EX/EY, EZ..FU — Minergie Kennzahlen
    (179, 180, 181, 202, "electric_energy", False),  # FW/FX, FY..GT — Elektrische Energie
    (204, 205, 206, 227, "strommodell", False),  # GV/GW, GX..HS
)

#: Sheet suffix -> value-kind variant of the ``Fläche-E`` family.
_FLAECHE_VARIANTS = (
    ("Fläche-E", "standard"),
    ("Fläche-ZW", "zielwert"),
    ("Fläche-Best", "bestand"),
    ("Fläche-L", "power"),
)

#: Provenance notes per per-category block kind (batch C).
_FLAECHE_BLOCK_NOTES = {
    "energy_standard": (
        "SIA 2024 Standardwerte (NGF/EBF, Faktor 0.8): Energiekennzahlen und "
        "Auslegungsparameter je Gebäudekategorie (Zeilen 3-47, Beschriftung "
        "Spalte AC, Einheit Spalte AD)"
    ),
    "energy_zielwert": (
        "SIA 2024 Zielwerte: Energiekennzahlen je Gebäudekategorie "
        "(Fläche-ZW, Beschriftung Spalte AC)"
    ),
    "energy_bestand": (
        "SIA 2024 Bestand: Energiekennzahlen je Gebäudekategorie "
        "(Fläche-Best, Beschriftung Spalte AC)"
    ),
    "energy_power": (
        "SIA 2024 Leistungswerte: Leistungsbedarf je Gebäudekategorie "
        "(Fläche-L, Beschriftung Spalte AC)"
    ),
    "sia3801_tab27": (
        "SIA 380/1 Tabelle 27 je Gebäudekategorie (Beschriftung Spalte BC, "
        "Einheit Spalte BD); Zeilen 15-17/35-38/43-45 sind die internen "
        "Wärmeeintragsleistungs-/Elektrizitätsbedarfs-/Reduktionsfaktor-Vergleiche "
        "desselben Blocks"
    ),
    "harmonized": (
        "Vorschlag harmonisierte Standardwerte, abgeleitet aus SIA 2024:2019 "
        "Anhang E (gerundete Werte; Beschriftung Spalte CB, Einheit Spalte CC)"
    ),
    "sia2040": (
        "SIA 2040:2017 je Gebäudekategorie (Beschriftung Spalte DB, Einheit "
        "Spalte DC); die SIA-390:2021-Untertabelle (Zeilen 21-31) ist nicht Teil "
        "dieser Tabelle"
    ),
    "weighted_energy": (
        "Minergie 2017, Gewichtete Energie (Beschriftung Spalte DZ, Einheit "
        "Spalte EA)"
    ),
    "minergie": (
        "Minergie 2017 Kennzahlen, umgerechnet in elektrische Energie "
        "(Beschriftung Spalte EX, Einheit Spalte EY)"
    ),
    "electric_energy": (
        "SIA 2024:2021, Elektrische Energie (Heizung/Warmwasser mit Wärmepumpe "
        "JAZ 4.0; Beschriftung Spalte FW, Einheit Spalte FX)"
    ),
    "strommodell": (
        "Strommodell für Zweckbauten: Geräte, Beleuchtung und Allgemeine "
        "Gebäudetechnik der Gebäude(haupt)nutzung (Beschriftung Spalte GV, "
        "Einheit Spalte GW)"
    ),
}

#: Row-1/row-label section headers (normalized) that start a redundant
#: sub-block inside the shared rows 3-47 (see ``_FLAECHE_BLOCKS``).
_SECTION_HEADER_LABELS = frozenset(
    {"sia 2024 zielwerte", "sia 2024 bestand", "sia 390:2021"}
)

#: Comparison sections of the rows 54-220 region (BC..BY columns, title in
#: BC): normalized title -> category-table kind.  The "Wärmeeinträge
#: Elektrizität" section (rows 60-63) has no matching kind and is not
#: extracted (documented scope decision).
_COMPARISON_KINDS = {
    "wärmebedarf warmwasser": "ww_demand",
    "aussenluft-volumenstrom": "ventilation_flow",
    "wärmeeinträge personen": "person_gain",
    "personenfläche": "person_area",
    "raumtemperatur": "room_temperature",
}

#: Unit-cell spelling variants (after superscript/dash normalization) mapped
#: onto registry symbols; the parenthesised flow spellings of the SIA 380/1
#: blocks differ from the registry entries.
_TABLE_UNIT_ALIASES = {
    "m3/(h·m2)": "m3/m2h",
    "m3/(m2h)": "m3/m2h",
    "m3/(h·P)": "m3/(Ph)",
}

_SUPERSCRIPT_TRANSLATION = str.maketrans(
    {
        "⁰": "0",
        "¹": "1",
        "²": "2",
        "³": "3",
        "⁴": "4",
        "⁵": "5",
        "⁶": "6",
        "⁷": "7",
        "⁸": "8",
        "⁹": "9",
        "⁻": "-",
    }
)


class DatasetExtractor:
    """Stage-0 extraction pipeline (assessment 7.1).

    Reads a **copy** of the source workbook deterministically (cell graph:
    values + formulas + cached results), checksums it, validates against the
    package JSON Schema and writes the canonical JSON package + manifest.

    Args:
        workbook_path: Path to a **copy** of ``2024_Raumdatenblätter_dfi_V221.xlsm``.
        output_dir: Directory the release package is written to.
        release_id: Release id of the produced package.
        extraction_tool_version: Version of this extraction tool (required).
        standard_version: Standard version tag of the extracted ``Volll_Lüft``
            block (V221 ships ``prSIA 2024-C1:2024``).

    Raises:
        FileNotFoundError: when the workbook copy does not exist (constructor).
    """

    def __init__(
        self,
        workbook_path: str,
        output_dir: str,
        release_id: str = "V221",
        *,
        extraction_tool_version: str,
        standard_version: str = "prSIA 2024-C1:2024",
    ) -> None:
        self.workbook_path = workbook_path
        self.output_dir = output_dir
        self.release_id = release_id
        self.extraction_tool_version = extraction_tool_version
        self.standard_version = standard_version
        if not os.path.isfile(workbook_path):
            raise FileNotFoundError(f"workbook copy not found: {workbook_path}")

    # -- pipeline -----------------------------------------------------------

    def extract(self) -> DatasetRelease:
        """Run the pipeline and return the release metadata of the written package.

        Raises:
            DatasetValidationError: unexpected sheet layout, missing required
                tables, unknown value kinds.
            OSError: unreadable copy.
        """
        try:
            # Lazily imported: openpyxl is part of the 'data' extra.
            import openpyxl  # type: ignore[import-untyped]
        except ImportError:
            raise ImportError(
                "openpyxl is required for workbook extraction; "
                "install it with `pip install 'energytools[data]'`"
            ) from None

        workbook_name = os.path.basename(self.workbook_path)
        source_checksum = self._checksum_file(self.workbook_path)

        # Two deterministic reads of the cell graph: formulas + cached results.
        wb_formulas = openpyxl.load_workbook(self.workbook_path, data_only=False, read_only=True)
        try:
            wb_values = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=True)
            try:
                self._check_layout(wb_values)
                package = self._build_package(
                    wb_values, wb_formulas, workbook_name, source_checksum
                )
            finally:
                wb_values.close()
        finally:
            wb_formulas.close()

        errors = _schema_errors(package)
        if errors:
            raise DatasetValidationError(
                f"extracted package for '{self.release_id}' fails schema validation",
                {"release_id": self.release_id, "errors": errors},
            )

        # Content validation: a corrupt/foreign extraction is never written.
        self._to_dataset(package)
        package["release"]["checksum_sha256"] = _content_checksum(package)
        release = _release_from_dict(package["release"])

        release_dir = Path(self.output_dir) / self.release_id
        release_dir.mkdir(parents=True, exist_ok=True)
        package_file = release_dir / "package.json"
        package_file.write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")
        schema_file = release_dir / "schema.json"
        schema_file.write_text(
            json.dumps(PACKAGE_SCHEMA, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return release

    # -- workbook parsing ---------------------------------------------------

    @staticmethod
    def _checksum_file(path: str) -> str:
        digest = hashlib.sha256()
        with open(path, "rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _check_layout(self, wb: Any) -> None:
        missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
        if missing:
            raise DatasetValidationError(
                f"workbook copy has unexpected sheet layout; missing required sheets: {missing}",
                {"release_id": self.release_id, "errors": [f"missing sheets: {missing}"]},
            )

    def _build_package(
        self, wb_values: Any, wb_formulas: Any, workbook_name: str, source_checksum: str
    ) -> dict:
        room_uses = self._extract_room_uses(wb_values["Eingabedaten"], wb_values["Begriffe"])
        by_nutzid = {room_use.nutzid: room_use for room_use in room_uses}
        by_code = {room_use.code: room_use for room_use in room_uses}
        catalog, selected_nutzid = self._extract_parameters(
            wb_values["Datenblatt"], wb_values["Begriffe"], by_nutzid
        )
        profiles, merged_catalog = self._extract_profiles(
            wb_values["Eingabedaten"], catalog, by_nutzid
        )
        schedules = self._extract_schedules(wb_values["Eingabedaten"], by_nutzid)
        room_use_inputs = self._extract_room_use_inputs(wb_values["Eingabedaten"], by_nutzid)
        sia2028_monthly = self._extract_sia2028(wb_values["Profile"])
        climate = self._extract_climate(
            wb_values["Winter_Auslegung"], wb_values["Monatswerte"], wb_values["Aug_Auslegung"]
        )
        full_load_hours = self._extract_full_load_hours(wb_values["Volll_Lüft"], by_code)
        qhc = self._extract_qhc(wb_values["Qhc_Klimastat"], by_code)
        hourly_profiles = self._extract_hourly_profiles(wb_values["Profile"])
        mappings, area_tables = self._extract_mappings_and_areas(
            wb_values["Fläche-E"], wb_values["GEPAMOD"], by_code
        )
        category_tables = self._extract_category_tables(
            wb_values["Fläche-E"],
            wb_values["Fläche-ZW"] if "Fläche-ZW" in wb_values.sheetnames else None,
            wb_values["Fläche-Best"] if "Fläche-Best" in wb_values.sheetnames else None,
            wb_values["Fläche-L"] if "Fläche-L" in wb_values.sheetnames else None,
            wb_values["GEPAMOD"],
        )
        sia3801, sia3801_coefficients = self._extract_sia3801(
            wb_values, by_nutzid, selected_nutzid, climate
        )

        return {
            "schema_version": "1.0",
            "source_checksum_sha256": source_checksum,
            "release": {
                "id": self.release_id,
                "edition": "SIA 2024",
                "publication_date": "2024-11-17",
                "checksum_sha256": "0" * 64,  # placeholder; replaced once content is final
                "source_workbook": workbook_name,
                "extraction_tool_version": self.extraction_tool_version,
                "supersedes": None,
                "changelog": [
                    {
                        "version": self.release_id,
                        "date": "2024-11-17",
                        "change": (
                            "Extracted by the energytools DatasetExtractor from the V221 "
                            "Raumdatenblätter workbook copy."
                        ),
                        "migration": None,
                    }
                ],
            },
            "room_uses": [room_use.as_dict() for room_use in room_uses],
            "parameters": [parameter.as_dict() for parameter in merged_catalog],
            "profiles": [
                {
                    "nutzid": nutzid,
                    "values": {
                        parameter_id: {
                            kind.value: value.as_dict() for kind, value in by_kind.items()
                        }
                        for parameter_id, by_kind in profile.values.items()
                    },
                }
                for nutzid, profile in sorted(profiles.items())
            ],
            "room_use_schedules": [schedule.as_dict() for schedule in schedules],
            "room_use_inputs": [inputs.as_dict() for inputs in room_use_inputs],
            "sia2028_monthly": (
                None if sia2028_monthly is None else sia2028_monthly.as_dict()
            ),
            "hourly_profiles": [profile.as_dict() for profile in hourly_profiles],
            "monthly_profiles": [
                profile.as_dict()
                for station in climate.stations
                for profile in station.monthly.values()
            ],
            "weekly_profiles": [],
            "climate": climate.as_dict(),
            "full_load_hours": full_load_hours.as_dict(),
            "qhc": qhc.as_dict(),
            "sia3801": [result.as_dict() for result in sia3801],
            "mappings": [mapping.as_dict() for mapping in mappings],
            "area_tables": [table.as_dict() for table in area_tables],
            "sia3801_coefficients": [coefficients.as_dict() for coefficients in sia3801_coefficients],
            "category_tables": [table.as_dict() for table in category_tables],
        }

    def _to_dataset(self, package: dict) -> Dataset:
        try:
            return Dataset.from_package_dict(package)
        except (ValueError, TypeError, KeyError) as exc:
            raise DatasetValidationError(
                f"extracted package for '{self.release_id}' has inconsistent content: {exc}",
                {"release_id": self.release_id, "errors": [str(exc)]},
            ) from exc

    # -- table extractors ---------------------------------------------------

    @staticmethod
    def _extract_room_uses(ws: Any, ws_begriffe: Any) -> list[RoomUse]:
        """``Eingabedaten!A9:C53`` (codes + German names), joined with ``Begriffe`` names."""
        name_map: dict[str, TrilingualText] = {}
        for row in ws_begriffe.iter_rows(min_row=134, max_row=178, min_col=2, max_col=5):
            if row[0].value is None:
                continue
            de = str(row[1].value or "").strip()
            if de:
                name_map[_normalize_label(de)] = TrilingualText(
                    de=de,
                    fr=str(row[2].value or "").strip(),
                    it=str(row[3].value or "").strip(),
                )
        room_uses = []
        for row_index, row in enumerate(
            ws.iter_rows(min_row=9, max_row=53, min_col=1, max_col=3), start=9
        ):
            code_raw = row[0].value
            de_name = str(row[2].value or "").strip()
            if code_raw is None or not de_name:
                continue
            code = normalize_room_use_code(str(code_raw).strip())
            category = int(code.split(".")[0])
            nutzid = row_index - 8
            name = name_map.get(_normalize_label(de_name), TrilingualText(de=de_name))
            room_uses.append(RoomUse(nutzid=nutzid, code=code, category=category, name=name))
        return room_uses

    def _extract_parameters(
        self, ws: Any, ws_begriffe: Any, by_nutzid: dict[int, RoomUse]
    ) -> tuple[list[Parameter], int]:
        """``Datenblatt`` rows 4-196 = the 193 data-sheet parameters.

        One catalog entry per non-empty row of the block (assessment 1.2).
        Labels come from column C and are joined with the ``Begriffe``
        dictionaries (rows 25-127: SIA Ziffern + DE/FR/IT; rows 183+: sheet
        labels) for trilingual labels and SIA clause ids.  Rows without a
        label stay in the catalog for a 1:1 row mapping: sub-case rows (F
        column carries a designator such as "Auslegung Heizung") get the
        label ``"<parent> (<designator>)"``, purely structural rows (section
        markers, table headers, spacing rows) an empty label.
        """
        ziffer_by_label: dict[str, str] = {}
        translations: dict[str, tuple[str, str]] = {}
        for row in ws_begriffe.iter_rows(min_row=25, max_row=300, min_col=2, max_col=5):
            ziffer = row[0].value
            label = row[1].value
            if ziffer is None and label is None:
                continue
            de = _normalize_label(str(label or ""))
            if not de:
                continue
            fr = str(row[2].value or "").strip()
            it = str(row[3].value or "").strip()
            if ziffer and label:
                ziffer_by_label.setdefault(de, str(ziffer).strip())
            translations.setdefault(de, (fr, it))

        nutzid_cell = _cell(ws, 1, 3)
        selected_nutzid = int(nutzid_cell) if nutzid_cell not in (None, "") else 1
        if selected_nutzid not in by_nutzid:
            raise DatasetValidationError(
                f"unexpected sheet layout: Datenblatt!C1 nutzid {selected_nutzid} is not in 1..45",
                {"release_id": self.release_id, "errors": [f"nutzid {selected_nutzid} unknown"]},
            )

        matrix = _read_matrix(ws, 4, 196, 1, 19)
        parameters: list[Parameter] = []
        used_ids: set[str] = set()
        used_symbols: set[str] = set()
        category = ""
        parent_label = ""
        for row_index, row in enumerate(matrix, start=4):
            # Every row of the block is a parameter row (rows 4-196 = 193
            # parameters, assessment 1.2); the empty row 112 (spacer) is kept
            # as a structural entry to preserve the 1:1 row mapping.
            section = row[0]
            if isinstance(section, str) and section.strip() in _SECTION_HEADERS:
                category = section.strip()
            designator = ""
            if isinstance(row[5], str) and row[5].strip():
                designator = row[5].strip()
            label = row[2]
            if isinstance(label, str) and label.strip():
                label = label.strip()
                if label != "berechnet!":  # computed-marker row: not a parent label
                    parent_label = label
            elif designator and parent_label:
                label = f"{parent_label} ({designator})"
            else:
                label = ""
            symbol = str(row[8] or "").strip()
            unit = _safe_unit(str(row[9] or ""))
            values_raw = (row[12], row[13], row[14])
            flags = tuple(_to_bool(row[idx]) for idx in (15, 16, 17, 18))

            parameter_id = self._parameter_id(
                label, symbol, ziffer_by_label, used_ids, used_symbols, row_index
            )
            value_kinds = tuple(
                kind
                for kind, raw in zip(
                    (ValueKind.STANDARD, ValueKind.ZIELWERT, ValueKind.BESTAND), values_raw
                )
                if _clean_cell_value(raw) is not None
            )
            data_type = _infer_data_type(values_raw)
            fr, it = translations.get(_normalize_label(label), ("", ""))
            parameters.append(
                Parameter(
                    id=parameter_id,
                    label=TrilingualText(de=label, fr=fr, it=it),
                    symbol=symbol,
                    unit=unit,
                    data_type=data_type,
                    category=category,
                    value_kinds=frozenset(value_kinds)
                    if value_kinds
                    else frozenset({ValueKind.STANDARD}),
                    export_flag=flags[0],
                    display_flag=flags[1],
                    internal_heat_flag=flags[2],
                    qhc_flag=flags[3],
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="Datenblatt",
                                range=f"A{row_index}:S{row_index}",
                            ),
                        )
                    ),
                )
            )
        return parameters, selected_nutzid

    def _parameter_id(
        self,
        label: str,
        symbol: str,
        ziffer_by_label: dict[str, str],
        used_ids: set[str],
        used_symbols: set[str],
        row_index: int | None = None,
    ) -> str:
        base = ziffer_by_label.get(_normalize_label(label))
        # Sub-case rows (symbol ends with ",C" / ",H", e.g. qi,des,C) get a
        # ".C" / ".H" suffix; the symbol base is the raw symbol without it.
        sub_suffix = ""
        symbol_base = symbol
        match = re.search(r",([CH])$", symbol)
        if match:
            sub_suffix = "." + match.group(1)
            symbol_base = symbol[:-2]
        if base is None:
            if symbol_base and symbol_base not in ("-", "–", "—") and symbol_base not in used_symbols:
                base = symbol_base
            elif label:
                base = _slugify(label)
            else:
                base = f"row-{row_index}" if row_index else "parameter"
        base = f"{base}{sub_suffix}"
        candidate = base
        counter = 2
        while candidate in used_ids:
            candidate = f"{base}-{counter}"
            counter += 1
        used_ids.add(candidate)
        if symbol and symbol not in ("-", "–", "—"):
            used_symbols.add(symbol)
        return candidate

    def _extract_profiles(
        self, ws: Any, catalog: list[Parameter], by_nutzid: dict[int, RoomUse]
    ) -> tuple[dict[int, RoomUseProfile], list[Parameter]]:
        """Per-room-use values from the ``Eingabedaten`` master matrix (rows 9-53).

        Only matrix columns whose name matches a catalog parameter are used
        (the workbook's own MATCH semantics); matrix-only columns (profile
        hour indices, monthly/weekly profile sections, comments, SIA 380/1
        system requirements) are not part of the canonical catalog.  When
        several columns share one label, the first matching catalog parameter
        that is not yet assigned takes the column (rows 136/137 = IDA codes /
        regulation names in the real workbook).

        Returns the profiles and the merged parameter catalog (the catalog is
        unchanged: no matrix-only parameters enter the canonical dataset).
        """
        catalog_by_label: dict[str, list[Parameter]] = {}
        for catalog_parameter in catalog:
            catalog_by_label.setdefault(
                _normalize_label(catalog_parameter.label.de), []
            ).append(catalog_parameter)
            # Sub-case parameters carry a " (<designator>)" label suffix; the
            # matrix columns use the bare base label.
            base_label = re.sub(r"\s+\([^)]*\)$", "", catalog_parameter.label.de)
            if base_label != catalog_parameter.label.de:
                catalog_by_label.setdefault(_normalize_label(base_label), []).append(
                    catalog_parameter
                )

        max_col = ws.max_column or 1
        matrix = _read_matrix(ws, 6, 53, 4, max_col)
        name_by_col: dict[int, str] = {}
        kind_by_col: dict[int, str] = {}
        unit_by_col: dict[int, str] = {}
        for col in range(4, max_col + 1):
            name = matrix[0][col - 4]
            if name is None or not str(name).strip():
                continue
            name_by_col[col] = str(name).strip()
            kind = matrix[1][col - 4]
            if kind is not None and str(kind).strip():
                kind_by_col[col] = str(kind).strip()
            unit = matrix[2][col - 4]
            if unit is not None and str(unit).strip():
                unit_by_col[col] = str(unit).strip()

        value_matrix = matrix[3:]  # rows 9..53
        if len(value_matrix) < 45:
            value_matrix.extend([None] * (max_col - 3) for _ in range(45 - len(value_matrix)))

        raw_values: dict[str, dict[int, dict[ValueKind, Any]]] = {}
        used_parameters: set[str] = set()
        col = 4  # column D
        while col <= max_col:
            name = name_by_col.get(col)
            if not name:
                col += 1
                continue
            kind_text = kind_by_col.get(col, "")  # row-7 kind of this column
            unit = _safe_unit(unit_by_col.get(col, "-"))
            matches = catalog_by_label.get(_normalize_label(name), [])

            kinds: tuple[ValueKind, ...]
            if (
                kind_text == "Standard"
                and col + 2 <= max_col
                and col + 1 not in name_by_col
                and col + 2 not in name_by_col
            ):
                group = [col, col + 1, col + 2]
                kinds = (ValueKind.STANDARD, ValueKind.ZIELWERT, ValueKind.BESTAND)
                col += 3
                sub_kind = None
            else:
                group = [col]
                kinds = (ValueKind.STANDARD,)
                sub_kind = {"Kühlfall": ".C", "Heizfall": ".H"}.get(kind_text)
                col += 1

            parameter = self._match_parameter(matches, sub_kind, name, unit, used_parameters)
            if parameter is None:
                continue  # matrix-only column: not part of the canonical catalog
            used_parameters.add(parameter.id)
            for nutzid in range(1, 46):
                row_values = value_matrix[nutzid - 1]
                by_kind = raw_values.setdefault(parameter.id, {}).setdefault(nutzid, {})
                for offset, kind in enumerate(kinds):
                    cell_value = row_values[group[offset] - 4]
                    cleaned = _clean_cell_value(cell_value)
                    if cleaned is not None:
                        by_kind.setdefault(kind, cleaned)

        catalog_dict = {parameter.id: parameter for parameter in catalog}
        profiles: dict[int, RoomUseProfile] = {}
        for nutzid, room_use in by_nutzid.items():
            values = {}
            for parameter_id, by_nutzid_values in raw_values.items():
                by_kind = by_nutzid_values.get(nutzid, {})
                values[parameter_id] = {
                    kind: ParameterValue(
                        parameter_id=parameter_id,
                        kind=kind,
                        value=value,
                        unit=catalog_dict[parameter_id].unit,
                    )
                    for kind, value in by_kind.items()
                }
            profiles[nutzid] = RoomUseProfile(
                room_use=room_use,
                values=values,
                parameter_catalog=catalog_dict,
            )
        return profiles, list(catalog_dict.values())

    def _extract_schedules(self, ws: Any, by_nutzid: dict[int, RoomUse]) -> list[RoomUseSchedule]:
        """``Eingabedaten`` rows 9-53: per-room-use time schedules.

        Matrix blocks (openpyxl column indices, 1-based): ``DP:EM`` person
        fraction (24 h), ``EN:FK`` device fraction (24 h), ``FM:FX`` annual
        profile (12 months), ``HC:HN`` previous annual profile (12 months),
        ``HS:HY`` weekly profile (7 days, day 1 = Saturday), ``FY`` rest days
        per week.  Empty cells are zero occupation (the workbook leaves them
        blank).  ``Nutzungstage pro Jahr`` (``365 - 52 * rest days``) and
        ``Jahresgleichzeitigkeit`` (mean of the 12 previous-month values) are
        derived from the workbook formulas, which carry no cached results.
        """
        # DP(120) .. HY(233); offsets: DP=0 EN=24 FM=49 FY=61 HC=91 HS=107
        matrix = _read_matrix(ws, 9, 53, 120, 233)
        schedules = []
        for nutzid in sorted(by_nutzid):
            row = matrix[nutzid - 1]
            person = tuple(_schedule_cell(row[col]) for col in range(24))
            device = tuple(_schedule_cell(row[col]) for col in range(24, 48))
            monthly = tuple(_schedule_cell(row[col]) for col in range(49, 61))
            rest_days = _schedule_cell(row[61])  # FY
            monthly_previous = tuple(_schedule_cell(row[col]) for col in range(91, 103))
            weekly = tuple(_schedule_cell(row[col]) for col in range(107, 114))
            schedules.append(
                RoomUseSchedule(
                    room_use_id=nutzid,
                    person_fraction=person,
                    device_fraction=device,
                    weekly_fraction=weekly,
                    monthly_fraction=monthly,
                    monthly_previous_fraction=monthly_previous,
                    rest_days_per_week=rest_days,
                    working_days_per_year=365.0 - 52.0 * rest_days,
                    annual_simultaneity=sum(monthly_previous) / 12.0,
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="Eingabedaten",
                                range="DP9:HY53",
                            ),
                        ),
                        note=(
                            "Personen-/Geräteprofil (Nutzungstag) 24 h, Wochenprofil 7 d, "
                            "Jahresprofil/Monatsprofil (bisher) 12 m, Ruhetage pro Woche; "
                            "Nutzungstage pro Jahr and Jahresgleichzeitigkeit derived "
                            "(365-52*rest_days, mean of Monatsprofil bisher)"
                        ),
                    ),
                )
            )
        return schedules

    def _extract_room_use_inputs(self, ws: Any, by_nutzid: dict[int, RoomUse]) -> list[RoomUseInputs]:
        """``Eingabedaten`` rows 9-53: design-input columns without catalog labels.

        Columns K, Y, Z, AF:AG, AK:AM, AT:AU, BM, BO, BW, CB, CX, DO, HB and
        the SIA 380/1 system-requirement block HZ:IE (see :class:`RoomUseInputs`
        for the exact mapping).  Missing cells stay ``None``.
        """
        # 1-based columns -> 0-based offsets within K(11)..IE(239)
        matrix = _read_matrix(ws, 9, 53, 11, 239)
        columns = {
            "fensteranteil": (0, float),  # K
            "solar_reduction_factor": (14, float),  # Y
            "shading_radiation_threshold": (15, float),  # Z
            "klimatisierung": (21, "x"),  # AF
            "klimatisierung_kategorie": (22, str),  # AG
            "schallschutz_key": (26, float),  # AK
            "schallschutz_geraete_db": (27, float),  # AL
            "schallschutz_nutzung_db": (28, float),  # AM
            "sensible_waerme_kuehlfall": (35, float),  # AT
            "sensible_waerme_heizfall": (36, float),  # AU
            "k0_korrektur": (54, float),  # BM
            "praesenzart": (56, str),  # BO
            "ida_kategorie": (64, str),  # BW
            "aussenluft_volumenstrom": (69, float),  # CB
            "cooling_necessity": (91, str),  # CX
            "tagesprofil_typ": (109, str),  # DO
            "monatsprofil_typ": (199, str),  # HB
            "qh_li0": (223, float),  # HZ
            "dqh_li": (224, float),  # IA
            "huellzahl": (227, float),  # ID
            "qh_lim": (228, float),  # IE
        }
        inputs = []
        for nutzid in sorted(by_nutzid):
            row = matrix[nutzid - 1]
            values: dict[str, object] = {}
            for name, (offset, kind) in columns.items():
                cell = row[offset] if offset < len(row) else None
                if cell is None:
                    continue
                if kind is float:
                    if isinstance(cell, (int, float)):
                        values[name] = float(cell)
                elif kind is str:
                    text = str(cell).strip()
                    if text:
                        values[name] = text
                elif kind == "x":
                    values[name] = str(cell).strip() != ""
            if not values:
                continue
            inputs.append(RoomUseInputs(room_use_id=nutzid, **values))
        return inputs

    def _extract_sia2028(self, ws_profile: Any) -> Sia2028Monthly | None:
        """``Profile!AS278:BD284``: the SIA 2028 monthly outdoor reference.

        Rows 279 (temperature °C), 282 (relative humidity %) and 284
        (room-temperature reference °C) of the embedded 12-month table;
        the saturation-pressure rows 280/281 are derived (Magnus) and not
        stored.
        """
        matrix = _read_matrix(ws_profile, 278, 284, 45, 56)  # AS..BD
        if len(matrix) < 3:
            return None
        temperature = [
            float(value)
            for value in matrix[1][:12]
            if isinstance(value, (int, float))
        ]
        humidity = [
            float(value)
            for value in matrix[4][:12]
            if isinstance(value, (int, float))
        ]
        room_temperature = [
            float(value)
            for value in matrix[6][:12]
            if isinstance(value, (int, float))
        ]
        if len(temperature) != 12 or len(humidity) != 12:
            return None
        return Sia2028Monthly(
            temperature=tuple(temperature),
            relative_humidity=tuple(humidity),
            room_temperature=tuple(room_temperature),
            provenance=Provenance(
                sources=(
                    SourceRef(
                        workbook=os.path.basename(self.workbook_path),
                        sheet="Profile",
                        range="AS278:BD284",
                    ),
                ),
                note="SIA 2028 Monatswerte (Aussentemperatur, relative Feuchte, Raumtemperatur)",
            ),
        )

    @staticmethod
    def _match_parameter(
        matches: list[Parameter],
        sub_kind: str | None,
        name: str,
        unit: str,
        used: set[str] | None = None,
    ) -> Parameter | None:
        """Pick the catalog parameter matching a matrix column (label + sub-case)."""
        used = used if used is not None else set()
        for parameter in matches:
            if parameter.id in used:
                continue
            if sub_kind is not None:
                if parameter.id.endswith(sub_kind):
                    return parameter
            elif not re.search(r",[CH]$", parameter.symbol) and not parameter.id.endswith(
                (".C", ".H")
            ):
                return parameter
        for parameter in matches:
            if parameter.id in used:
                continue
            return parameter
        if matches:
            return matches[0]
        return None

    def _extract_climate(self, ws_winter: Any, ws_monatswerte: Any, ws_aug: Any) -> ClimateData:
        stations = []
        monthly_blocks: dict[str, dict[str, MonthlyProfile]] = {}
        # Monatswerte: 11-row blocks per station starting at rows 1, 12, 23, ...
        # (row 1 is a formula-driven preview of the selected station; the
        # literal blocks repeat the same data).  The air temperature row is
        # the row BELOW the block header.  Block names are joined to the
        # Winter_Auslegung stations by a normalized key (the workbook spells
        # station 5 "Bern Liebefeld" in Monatswerte but "Bern-Liebefeld" in
        # Winter_Auslegung).
        monthly = _read_matrix(ws_monatswerte, 1, 450, 1, 23)
        for block in range(0, 450, 11):
            name = str(monthly[block][0] or "").strip()
            if not name:
                continue
            monthly_block = monthly_blocks.setdefault(_normalize_station_name(name), {})
            values = [_month_value(value) for value in monthly[block + 1][11:23]]
            if all(value is None for value in values):
                continue  # header-only block
            monthly_block["t_aussen"] = MonthlyProfile(
                id="t_aussen",
                values=tuple(value if value is not None else 0.0 for value in values),
                unit="°C",
            )
            for key, row_offset in (
                ("radiation_horizontal", 2),
                ("radiation_east", 3),
                ("radiation_south", 4),
                ("radiation_west", 5),
                ("radiation_north", 6),
                ("precipitation", 7),
                ("mixing_ratio", 8),
                ("absolute_humidity", 9),
            ):
                values_row = monthly[block + row_offset]
                values = [_month_value(value) for value in values_row[11:23]]
                if any(value is not None for value in values):
                    monthly_block[key] = MonthlyProfile(
                        id=key,
                        values=tuple(value if value is not None else 0.0 for value in values),
                        unit=_MONTHLY_BLOCK_UNITS[key],
                    )

        winter = _read_matrix(ws_winter, 1, 44, 1, 35)  # A..AI
        design_day_blocks: dict[str, list[DesignDaySeries]] = {}
        if ws_aug is not None:
            design_day_blocks = self._extract_design_days(ws_aug)
        for index, row in enumerate(winter[4:44], start=5):  # sheet rows 5..44
            name = str(row[0] or "").strip()
            if not name:
                continue
            station_id = index - 4
            winter_design: dict[str, Quantity] = {}
            for key, column, unit in (
                ("t_a", 7, "°C"),  # H column: design temperature
                ("t_heating", 5, "°C"),  # F column: Heizung
                ("t_ventilation", 6, "°C"),  # G column: Lüftung
                ("radiation", 9, "W/m2"),  # J column: horizontal
                ("radiation_east", 11, "W/m2"),  # L..O: kalt E/S/W/N
                ("radiation_south", 12, "W/m2"),
                ("radiation_west", 13, "W/m2"),
                ("radiation_north", 14, "W/m2"),
                ("wind_speed", 15, "m/s"),  # P: kalt wind speed
                ("elevation", 2, "m"),  # C: Höhe m ü.M.
                ("trub_temperature", 19, "°C"),  # T: trüb design temperature
                ("trub_radiation", 21, "W/m2"),  # V: trüb horizontal
                ("trub_radiation_east", 23, "W/m2"),  # X..AA: trüb E/S/W/N
                ("trub_radiation_south", 24, "W/m2"),
                ("trub_radiation_west", 25, "W/m2"),
                ("trub_radiation_north", 26, "W/m2"),
                ("trub_wind_speed", 27, "m/s"),  # AB: trüb wind speed
                ("t_min_1h", 31, "°C"),  # AF: minimale 1-Stunden-Temperatur
                ("humidity_ratio_min", 33, "g/kg"),  # AH: Feuchtegehalt
                ("t_at_min", 34, "°C"),  # AI: zugehörige Temperatur
            ):
                value = row[column]
                if isinstance(value, (int, float)):
                    winter_design[key] = Quantity(float(value), unit)
            hdd_value = row[4]  # E column: Heizgradtage
            hdd = Quantity(float(hdd_value), "K·d") if isinstance(hdd_value, (int, float)) else None
            stations.append(
                ClimateStation(
                    id=station_id,
                    name=TrilingualText(de=name),
                    winter_design=winter_design,
                    summer_design={},
                    monthly=monthly_blocks.get(_normalize_station_name(name), {}),
                    hdd=hdd,
                    canton=str(row[1] or "").strip() or None,  # B: Kanton
                    wind_direction=str(row[17] or "").strip() or None,  # R: kalt 30°-Sektor
                    trub_wind_direction=str(row[29] or "").strip() or None,  # AD: trüb 30°-Sektor
                    design_days=tuple(design_day_blocks.get(_normalize_station_name(name), ())),
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="Winter_Auslegung",
                                range=f"A{index}:AI{index}",
                            ),
                        )
                    ),
                )
            )
        return ClimateData(
            version="meteoschweiz-2024", stations=tuple(stations), source="MeteoSchweiz"
        )

    def _extract_design_days(self, ws_aug: Any) -> dict[str, list[DesignDaySeries]]:
        """``Aug_Auslegung``: June + August 96-hour design blocks per station.

        Row 1 carries the 40 station names starting at column K, one
        3-column group each (Aussenlufttemp. in 0.1 °C, rel. Luftfeuchte %,
        Globalstrahlung W/m²).  Rows 4-99 are the June block, rows 100-195
        the August block (4 x 24 h each).
        """
        blocks: dict[str, list[DesignDaySeries]] = {}
        matrix = _read_matrix(ws_aug, 1, 195, 1, ws_aug.max_column or 1)
        station_names: list[str] = []
        for column in range(10, (ws_aug.max_column or 1) - 1, 3):  # K..DZ groups
            name = str(matrix[0][column] or "").strip()
            if not name:
                break
            station_names.append(name)
        for month, start_row, end_row in ((6, 4, 100), (8, 100, 196)):
            for station_index, name in enumerate(station_names):
                base = 10 + 3 * station_index
                if base + 2 >= (ws_aug.max_column or 1):
                    continue
                temperature = []
                relative_humidity = []
                radiation = []
                for row_index in range(start_row, end_row):
                    row_values = matrix[row_index - 1]
                    t_value = row_values[base] if base < len(row_values) else None
                    rh_value = row_values[base + 1] if base + 1 < len(row_values) else None
                    r_value = row_values[base + 2] if base + 2 < len(row_values) else None
                    temperature.append(
                        float(t_value) / 10.0
                        if isinstance(t_value, (int, float))
                        else 0.0
                    )
                    relative_humidity.append(
                        float(rh_value) if isinstance(rh_value, (int, float)) else 0.0
                    )
                    radiation.append(
                        float(r_value) if isinstance(r_value, (int, float)) else 0.0
                    )
                blocks.setdefault(_normalize_station_name(name), []).append(
                    DesignDaySeries(
                        month=month,
                        temperature=tuple(temperature),
                        relative_humidity=tuple(relative_humidity),
                        radiation=tuple(radiation),
                        provenance=Provenance(
                            sources=(
                                SourceRef(
                                    workbook=os.path.basename(self.workbook_path),
                                    sheet="Aug_Auslegung",
                                    range=f"K{start_row}:{_column_name(base + 3)}{end_row - 1}",
                                ),
                            )
                        ),
                    )
                )
        return blocks

    def _extract_full_load_hours(self, ws: Any, by_code: dict[str, RoomUse]) -> FullLoadHoursTable:
        rows: dict[tuple[int, str, str], float] = {}
        electrical: dict[tuple[int, str, str], float] = {}
        stage_hours: dict[tuple[int, str, float, str], float] = {}
        regulations = ("1-stufig", "2-stufig", "stufenlos")
        # D/F/J: Volllaststunden Volumenstrom; E/I/Q: Volllaststunden
        # elektrische Energie; G/H: Betriebsstunden 2-stufig 67/100 %;
        # K..P: Betriebsstunden stufenlos 25/40/50/60/80/100 %.
        electrical_columns = {"1-stufig": 4, "2-stufig": 8, "stufenlos": 16}  # 1-based E/I/Q
        stage_columns = {
            "2-stufig": ((67.0, 6), (100.0, 7)),  # G/H
            "stufenlos": ((25.0, 10), (40.0, 11), (50.0, 12), (60.0, 13), (80.0, 14), (100.0, 15)),
        }
        for row in ws.iter_rows(min_row=7, max_row=51, min_col=1, max_col=17):
            code = str(row[0].value or "").strip()
            room_use = by_code.get(normalize_room_use_code(code))
            if room_use is None:
                continue
            for regulation, column_index in zip(regulations, (3, 5, 9)):
                value = row[column_index].value
                if value is not None and isinstance(value, (int, float)):
                    rows[(room_use.nutzid, regulation, self.standard_version)] = float(value)
                electrical_value = row[electrical_columns[regulation]].value
                if electrical_value is not None and isinstance(electrical_value, (int, float)):
                    electrical[
                        (room_use.nutzid, regulation, self.standard_version)
                    ] = float(electrical_value)
                for stage, column_index in stage_columns.get(regulation, ()):
                    stage_value = row[column_index].value
                    if stage_value is not None and isinstance(stage_value, (int, float)):
                        stage_hours[
                            (room_use.nutzid, regulation, stage, self.standard_version)
                        ] = float(stage_value)
        return FullLoadHoursTable(
            rows=rows,
            electrical=electrical,
            stage_hours=stage_hours,
            standard_versions=frozenset({self.standard_version}),
            regulations=frozenset(regulations),
            default_standard_version=self.standard_version,
            provenance=Provenance(
                sources=(
                    SourceRef(
                        workbook=os.path.basename(self.workbook_path),
                        sheet="Volll_Lüft",
                        range="A7:Q51",
                    ),
                ),
                note=(
                    "Volllaststunden Volumenstrom (D/F/J) und elektrische Energie (E/I/Q), "
                    "Betriebsstunden Volumenstrom je Stufe (G/H, K..P); "
                    f"standard version {self.standard_version}"
                ),
            ),
        )

    def _extract_qhc(self, ws: Any, by_code: dict[str, RoomUse]) -> QhcTable:
        """``Qhc_Klimastat``: 12 columns per station block, E/I/M = annual cooling per kind.

        Each 12-column station block carries four metrics per value kind
        (offsets 0..3 within the kind): cooling power (W/m²), annual cooling
        energy (kWh/m²a), heating design load (Norm-Heizlast, W/m²) and
        annual heating energy (kWh/m²a).
        """
        rows: dict[tuple[int, int, ValueKind], float] = {}
        cooling_power: dict[tuple[int, int, ValueKind], float] = {}
        heating_load: dict[tuple[int, int, ValueKind], float] = {}
        heating_energy: dict[tuple[int, int, ValueKind], float] = {}
        matrix = _read_matrix(ws, 1, 51, 1, ws.max_column or 1)
        station_start = 4  # column D
        station_id = 1
        while station_start <= (ws.max_column or 1):
            station_name_cell = matrix[2][station_start - 1]
            if station_name_cell is None:
                break
            for kind, offset in (
                (ValueKind.STANDARD, 0),
                (ValueKind.ZIELWERT, 4),
                (ValueKind.BESTAND, 8),
            ):
                metric_columns = (station_start + offset, station_start + offset + 1, station_start + offset + 2, station_start + offset + 3)
                for row_index in range(7, 52):  # sheet rows 7..51
                    code = str(matrix[row_index - 1][0] or "").strip()
                    room_use = by_code.get(normalize_room_use_code(code))
                    if room_use is None:
                        continue
                    row_values = matrix[row_index - 1]
                    for metric, column, target in (
                        ("power", 0, cooling_power),
                        ("cooling", 1, rows),
                        ("load", 2, heating_load),
                        ("heating", 3, heating_energy),
                    ):
                        value_column = metric_columns[column]
                        if value_column - 1 >= len(row_values):
                            continue  # block extends beyond the sheet (short sheets)
                        value_cell = row_values[value_column - 1]
                        if isinstance(value_cell, (int, float)):
                            target[(room_use.nutzid, station_id, kind)] = float(value_cell)
            station_start += 12
            station_id += 1
        return QhcTable(
            rows={key: Quantity(value, "kWh/m2") for key, value in rows.items()},
            cooling_power={key: Quantity(value, "W/m2") for key, value in cooling_power.items()},
            heating_load={key: Quantity(value, "W/m2") for key, value in heating_load.items()},
            heating_energy={key: Quantity(value, "kWh/m2") for key, value in heating_energy.items()},
            provenance=Provenance(
                sources=(
                    SourceRef(
                        workbook=os.path.basename(self.workbook_path),
                        sheet="Qhc_Klimastat",
                        range="A7:M51",
                    ),
                ),
                note=(
                    "Vier Metriken je Wertkategorie: Klimakälteleistungsbedarf (W/m2), "
                    "Jährlicher Klimakältebedarf (kWh/m2), Norm-Heizlast (W/m2), "
                    "Jährlicher Heizwärmebedarf (kWh/m2), D/H/L + E/I/M + F/J/N + G/K/O "
                    "Spalten je Station"
                ),
            ),
        )

    def _extract_hourly_profiles(self, ws: Any) -> list[HourlyProfile]:
        """``Profile`` rows 63-86: person / device / ventilation hour fractions."""
        columns = (
            ("personen_werktag", "person", 1),
            ("geraete_werktag", "device", 3),
            ("geraete_freier_tag", "device", 4),
            ("lueftung_einstufig", "ventilation", 5),
            ("lueftung_zweistufig", "ventilation", 6),
            ("lueftung_stufenlos", "ventilation", 8),
        )
        matrix = _read_matrix(ws, 63, 86, 1, 22)
        profiles = []
        for profile_id, profile_type, column_index in columns + (
            ("beleuchtung_sommer", "lighting", 19),  # T column: Beleuchtung Sommer
            ("beleuchtung_jahr", "lighting", 21),  # V column: Beleuchtung Jahr
        ):
            values = [
                float(row[column_index]) if isinstance(row[column_index], (int, float)) else 0.0
                for row in matrix
            ]
            profiles.append(
                HourlyProfile(
                    id=profile_id,
                    profile_type=profile_type,
                    values=tuple(values),
                    unit="%",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="Profile",
                                range="A63:V86",
                            ),
                        ),
                        note=(
                            "Tagesprofile (63-86); Beleuchtung Sommer/Jahr sind die "
                            "Tageslicht-regelungsabhängigen Beleuchtungsprofile des "
                            "ausgewählten Raumnutzungs-Sichtblatts"
                        ),
                    ),
                )
            )
        return profiles

    def _extract_mappings_and_areas(
        self, ws_flaeche: Any, ws_gepamod: Any, by_code: dict[str, RoomUse]
    ) -> tuple[list[BuildingCategoryMapping], list[AreaTable]]:
        """``Fläche-E`` rows 3-49: category columns (row 1) x room-use codes (column A).

        Only the first category block of row 1 is read (columns D..Y in the
        real workbook; the later blocks hold variant metrics of other
        standards).  Category names come from ``GEPAMOD`` row 2 per category
        column of row 1.
        """
        gepamod = _read_matrix(ws_gepamod, 1, 2, 1, ws_gepamod.max_column or 41)
        category_names: dict[str, str] = {}
        for col in range(len(gepamod[0])):
            code = gepamod[0][col]
            name = gepamod[1][col]
            if code is None or name is None:
                continue
            code = str(code).strip()
            name = str(name).strip()
            if code and name:
                category_names.setdefault(code, name)

        flaeche = _read_matrix(ws_flaeche, 1, 49, 1, ws_flaeche.max_column or 1)
        categories: list[str] = []
        for cell in flaeche[0][3:]:  # first block: consecutive row-1 cells from column D
            if cell is None or not str(cell).strip():
                break
            categories.append(str(cell).strip())

        rows: dict[str, dict[str, float]] = {category: {} for category in categories}
        for row_index, row in enumerate(flaeche[2:49], start=3):  # sheet rows 3..49
            code_raw = row[0]
            if code_raw is None:
                continue
            code = normalize_room_use_code(str(code_raw).strip())
            if code not in by_code:
                continue
            for column_index, category in enumerate(categories, start=4):
                value_cell = row[column_index - 1]
                if isinstance(value_cell, (int, float)):
                    rows[category][code] = float(value_cell)

        mappings = [
            BuildingCategoryMapping(
                sia3801_category=category,
                room_use_codes=frozenset(codes),
                name=TrilingualText(de=category_names.get(category, ""))
                if category_names.get(category)
                else None,
                provenance=Provenance(
                    sources=(
                        SourceRef(
                            workbook=os.path.basename(self.workbook_path),
                            sheet="GEPAMOD",
                            range="A1:L2",
                        ),
                    )
                ),
            )
            for category, codes in rows.items()
            if codes
        ]
        area_table = AreaTable(
            kind=ValueKind.STANDARD,
            rows={
                category: {code: Quantity(value, "%") for code, value in codes.items()}
                for category, codes in rows.items()
                if codes
            },
            provenance=Provenance(
                sources=(
                    SourceRef(
                        workbook=os.path.basename(self.workbook_path),
                        sheet="Fläche-E",
                        range="A3:C49",
                    ),
                ),
                note="Anteil der Raumfläche an der Gebäudekategorie (%)",
            ),
        )
        return mappings, [area_table]

    def _extract_category_tables(
        self,
        ws_flaeche_e: Any,
        ws_flaeche_zw: Any | None,
        ws_flaeche_best: Any | None,
        ws_flaeche_l: Any | None,
        ws_gepamod: Any,
    ) -> list[CategoryTable]:
        """Per-category reference tables of the Fläche family + GEPAMOD (batch C).

        Reads the cached values of the per-category blocks (rows 3-47 of each
        Fläche sheet, one metric per row with a private label column) and the
        SIA 380/1 vs SIA 2024 comparison rows (rows 54-220), plus the GEPAMOD
        subcategory / EBF / end-energy rows.  Formula cells without a cached
        value are skipped; missing sheets (synthetic workbooks) are skipped.
        """
        tables: list[CategoryTable] = []
        for sheet_name, variant in _FLAECHE_VARIANTS:
            ws = {
                "Fläche-E": ws_flaeche_e,
                "Fläche-ZW": ws_flaeche_zw,
                "Fläche-Best": ws_flaeche_best,
                "Fläche-L": ws_flaeche_l,
            }[sheet_name]
            if ws is None:
                continue
            tables.extend(self._extract_flaeche_category_tables(ws, sheet_name, variant))
        tables.extend(self._extract_gepamod_tables(ws_gepamod))
        return tables

    def _extract_flaeche_category_tables(
        self, ws: Any, sheet_name: str, variant: str
    ) -> list[CategoryTable]:
        """The per-category blocks and comparison rows of one Fläche sheet."""
        matrix = _read_matrix(ws, 1, 220, 1, ws.max_column or 1)
        tables: list[CategoryTable] = []
        for label_col, unit_col, first_cat, last_cat, kind, stop_at_header in _FLAECHE_BLOCKS:
            if kind is None:
                kind = f"energy_{variant}"
            categories, metric_rows = _category_block_rows(
                matrix, label_col, unit_col, first_cat, last_cat, stop_at_header
            )
            if not categories or not metric_rows:
                continue
            rows_map = {}
            used_labels: set[str] = set()
            units: set[str] = set()
            for row_number, label, unit, values in metric_rows:
                if label in used_labels:
                    label = f"{label} (row {row_number})"
                used_labels.add(label)
                units.add(unit)
                for col, value in values:
                    rows_map[(categories[col], label)] = Quantity(float(value), unit)
            tables.append(
                CategoryTable(
                    kind=kind,
                    variant=variant,
                    rows=rows_map,
                    unit=next(iter(units)) if len(units) == 1 else "-",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet=sheet_name,
                                range=(
                                    f"{_column_name(label_col)}3:"
                                    f"{_column_name(last_cat)}47"
                                ),
                            ),
                        ),
                        note=_FLAECHE_BLOCK_NOTES.get(kind),
                    ),
                )
            )
        tables.extend(self._extract_flaeche_comparisons(matrix, sheet_name))
        return tables

    def _extract_flaeche_comparisons(self, matrix: list[list[Any]], sheet_name: str) -> list[CategoryTable]:
        """The SIA 380/1 vs SIA 2024 comparison rows (rows 54-220, BC..BY).

        Sections start with a title row (BC label, no values); the per-person
        air-flow rows (unit ``m3/(Ph)``) carry no title and are classified by
        their unit.  The "Wärmeeinträge Elektrizität" section (rows 60-63) has
        no matching kind and is skipped.
        """
        label_col, unit_col, first_cat, last_cat = 55, 56, 57, 78
        categories = _category_codes(matrix, first_cat, last_cat)
        if not categories:
            return []
        sections: list[tuple[str, int, int, list]] = []
        current_kind: str | None = None
        current_rows: list = []
        current_start = 0
        current_unit: str | None = None

        def flush() -> None:
            nonlocal current_kind, current_rows, current_start, current_unit
            if current_kind and current_rows:
                sections.append(
                    (current_kind, current_start, current_rows[-1][0], current_rows)
                )
            current_kind, current_rows, current_start, current_unit = None, [], 0, None

        for row_number in range(54, 221):
            row = matrix[row_number - 1] if row_number - 1 < len(matrix) else [None] * 1
            label = row[label_col - 1] if label_col - 1 < len(row) else None
            if label is None or not str(label).strip():
                continue
            values = [
                (col, row[col - 1])
                for col in range(first_cat, last_cat + 1)
                if col - 1 < len(row) and isinstance(row[col - 1], (int, float))
            ]
            unit = _category_table_unit(row[unit_col - 1] if unit_col - 1 < len(row) else None)
            if not values:
                # title row: a section boundary
                flush()
                current_kind = _COMPARISON_KINDS.get(_normalize_label(str(label).strip()))
                current_start = row_number
                continue
            if current_kind is None:
                # value rows without a preceding title (per-person air flow):
                # classified by their unit
                if unit == "m3/(Ph)":
                    current_kind = "ventilation_flow"
                    current_start = row_number
                else:
                    continue
            if current_rows and unit != current_rows[-1][2]:
                # a unit change separates two unnamed sections (the per-area and
                # per-person air-flow rows share no title rows)
                flush()
                if unit == "m3/(Ph)":
                    current_kind = "ventilation_flow"
                    current_start = row_number
                else:
                    continue
            current_rows.append((row_number, str(label).strip(), unit, values))
        flush()

        tables: list[CategoryTable] = []
        for kind, start, end, metric_rows in sections:
            rows_map = {}
            used_labels: set[str] = set()
            units: set[str] = set()
            for row_number, label, unit, values in metric_rows:
                if label in used_labels:
                    label = f"{label} (row {row_number})"
                used_labels.add(label)
                units.add(unit)
                for col, value in values:
                    rows_map[(categories[col], label)] = Quantity(float(value), unit)
            tables.append(
                CategoryTable(
                    kind=kind,
                    variant="reference",
                    rows=rows_map,
                    unit=next(iter(units)) if len(units) == 1 else "-",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet=sheet_name,
                                range=f"BC{start}:BY{end}",
                            ),
                        ),
                        note=(
                            "SIA 380/1 vs SIA 2024 Vergleichswerte je Gebäudekategorie "
                            "(BC..BY, Zeilen 54-220); Raumtemperatur-Vergleich führt die "
                            "Arbeitsmappe-Einheit 'm2' (Arbeitsmappen-Quirk)"
                        ),
                    ),
                )
            )
        return tables

    def _extract_gepamod_tables(self, ws_gepamod: Any) -> list[CategoryTable]:
        """GEPAMOD subcategory / subsubsector / EBF / end-energy tables (rows 3-14)."""
        if ws_gepamod is None:
            return []
        matrix = _read_matrix(ws_gepamod, 1, 14, 1, 41)
        gepamod_categories = _category_codes(matrix, 4, 17)  # D..Q (SIA 380/1 subcodes)
        sia2024_categories = _category_codes(matrix, 21, 41)  # U..AO (SIA 2024 codes)
        tables: list[CategoryTable] = []

        # rows 3-4: Unterkategorie labels + Subsubsektor id lists (strings)
        subcategory_rows: dict[tuple[str, str], Quantity] = {}
        for row_number, metric in (
            (3, "Unterkategorie"),
            (4, "Subsubsektor Gebaeudeparkmodell"),
        ):
            row = matrix[row_number - 1]
            for col, code in gepamod_categories.items():
                cell = row[col - 1] if col - 1 < len(row) else None
                if cell is not None and str(cell).strip():
                    subcategory_rows[(code, metric)] = Quantity(str(cell).strip(), "-")
        if subcategory_rows:
            tables.append(
                CategoryTable(
                    kind="gepamod_subcategory",
                    variant="reference",
                    rows=subcategory_rows,
                    unit="-",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="GEPAMOD",
                                range="A3:Q4",
                            ),
                        ),
                        note=(
                            "Unterkategorie-Bezeichnungen und Subsubsektor-IDs des "
                            "Gebäudeparkmodells je SIA-380/1-Subkategorie (D:Q)"
                        ),
                    ),
                )
            )

        # row 5: EBF (Modell 2010, 1000 m2)
        row = matrix[4]
        ebf_rows: dict[tuple[str, str], Quantity] = {}
        for col, code in gepamod_categories.items():
            cell = row[col - 1] if col - 1 < len(row) else None
            if isinstance(cell, (int, float)):
                ebf_rows[(code, "EBF")] = Quantity(float(cell), "1000m2")
        if ebf_rows:
            tables.append(
                CategoryTable(
                    kind="gepamod_ebf",
                    variant="reference",
                    rows=ebf_rows,
                    unit="1000m2",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="GEPAMOD",
                                range="A5:Q5",
                            ),
                        ),
                        note="Energiebezugsfläche EBF, Modell 2010 (1000 m2) je Subkategorie",
                    ),
                )
            )

        # rows 6-14: Neubau-2010 Endenergie / Nutzenergie intensities.  The
        # GEPAMOD columns (D:Q) carry their own row labels (column B); the SIA
        # 2024 side columns (U:AO) use the column-S labels, which are offset
        # from column B in rows 9/10 (workbook layout), so the two sides are
        # extracted as separate tables.
        def _end_energy_rows(
            label_col: int, categories: dict[int, str], first_cat: int, last_cat: int
        ) -> dict[tuple[str, str], Quantity]:
            result: dict[tuple[str, str], Quantity] = {}
            section = "Endenergie"
            for row_number in range(6, 15):
                row = matrix[row_number - 1]
                section_cell = row[0] if len(row) > 0 else None
                if section_cell is not None and str(section_cell).strip():
                    section = (
                        re.sub(r"^Neubau\s+2010\s+", "", str(section_cell).strip()) or section
                    )
                label_cell = row[label_col - 1] if label_col - 1 < len(row) else None
                if label_cell is None or not str(label_cell).strip():
                    continue
                metric = f"{section} / {str(label_cell).strip()}"
                for col, code in categories.items():
                    cell = row[col - 1] if col - 1 < len(row) else None
                    if isinstance(cell, (int, float)):
                        result[(code, metric)] = Quantity(float(cell), "kWh/m2")
            return result

        gepamod_end = _end_energy_rows(2, gepamod_categories, 4, 17)  # column B, D..Q
        if gepamod_end:
            tables.append(
                CategoryTable(
                    kind="gepamod_end_energy",
                    variant="reference",
                    rows=gepamod_end,
                    unit="kWh/m2",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="GEPAMOD",
                                range="A6:Q14",
                            ),
                        ),
                        note=(
                            "Neubau-2010 Endenergie/Nutzenergie je Subkategorie "
                            "(GEPAMOD Modell 2010, Spalten D:Q; Zeilenbeschriftung Spalte B)"
                        ),
                    ),
                )
            )
        sia2024_end = _end_energy_rows(19, sia2024_categories, 21, 41)  # column S, U..AO
        if sia2024_end:
            tables.append(
                CategoryTable(
                    kind="gepamod_end_energy",
                    variant="reference",
                    rows=sia2024_end,
                    unit="kWh/m2",
                    provenance=Provenance(
                        sources=(
                            SourceRef(
                                workbook=os.path.basename(self.workbook_path),
                                sheet="GEPAMOD",
                                range="A6:AO14",
                            ),
                        ),
                        note=(
                            "Neubau-2010 Endenergie/Nutzenergie je SIA-2024-Kategorie "
                            "(Spalten U:AO; Zeilenbeschriftung Spalte S)"
                        ),
                    ),
                )
            )
        return tables

    def _extract_sia3801(
        self, wb: Any, by_nutzid: dict[int, RoomUse], selected_nutzid: int, climate: ClimateData
    ) -> tuple[list[Sia3801Result], list[Sia3801Coefficients]]:
        """``SIA 380-1`` family: Qh/Qc per value kind + the coefficient block.

        The four sheets are one calculation with a variant axis.  Qh (kWh/m²a)
        lives at P134/P166/P196 (yearly row ``P133/3.6``), except the ``_Qc_EN``
        sheet where those rows are absent — there the yearly totals P133/P165/
        P195 (MJ/m²a) are converted.  The cooling demand ``Qc`` (kWh/m²a) lives
        at P137/P167/P197.  Each sheet names its own climate station in B68
        (formula ``Monatswerte!A1`` on the EN sheets, resolved through
        ``Eigene Nutzung!G1`` -> ``Winter_Auslegung`` row).  The coefficient
        block (rows 42-63, column D) is read per variant; ``Thetai`` (r58) is
        formula-derived and left to the engine.
        """
        variants = {
            "SIA 380-1": "de",
            "SIA 380-1_EN": "en",
            "SIA 380-1_Qc": "de+qc",
            "SIA 380-1_Qc_EN": "en+qc",
        }
        # The EN sheets reference the selected station through
        # Eigene Nutzung!G1 -> Winter_Auslegung rows 5..44.
        station_by_name = {station.name.de: station.id for station in climate.stations}
        eigene_g1 = _cell(wb["Eigene Nutzung"], 1, 7)  # G1: selected station index
        winter_matrix = _read_matrix(wb["Winter_Auslegung"], 5, 44, 1, 1)
        selected_station_name = None
        if isinstance(eigene_g1, (int, float)) and 1 <= int(eigene_g1) <= 40:
            cell = winter_matrix[int(eigene_g1) - 1][0]
            if cell:
                selected_station_name = str(cell).strip()

        results: list[Sia3801Result] = []
        coefficients: list[Sia3801Coefficients] = []
        for sheet_name, variant in variants.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            matrix = _read_matrix(ws, 1, 197, 1, 16)
            station_name = str(matrix[67][1] or "").strip()  # B68
            if station_name in station_by_name:
                station_id = station_by_name[station_name]
            elif selected_station_name in station_by_name:
                # EN/Qc_EN B68 is the formula Monatswerte!A1 (no cached text)
                station_id = station_by_name[selected_station_name]
            else:
                station_id = climate.stations[0].id
            for kind, qh_row, fallback_row, qc_row in (
                (ValueKind.STANDARD, 134, 133, 137),
                (ValueKind.ZIELWERT, 166, 165, 167),
                (ValueKind.BESTAND, 196, 195, 197),
            ):
                values: dict[str, Quantity] = {}
                value_cell = matrix[qh_row - 1][15]  # P134/P166/P196
                if not isinstance(value_cell, (int, float)):
                    # Qc_EN lacks those rows; convert the yearly total P133/3.6
                    total_cell = matrix[fallback_row - 1][15]
                    if isinstance(total_cell, (int, float)):
                        value_cell = float(total_cell) / 3.6
                if isinstance(value_cell, (int, float)):
                    values["Qh"] = Quantity(float(value_cell), "kWh/m2a")
                qc_cell = matrix[qc_row - 1][15]  # P137/P167/P197 (MJ/m2a)
                if isinstance(qc_cell, (int, float)):
                    values["Qc"] = Quantity(float(qc_cell) / 3.6, "kWh/m2a")
                if not values:
                    continue
                results.append(
                    Sia3801Result(
                        room_use_id=selected_nutzid,
                        station_id=station_id,
                        kind=kind,
                        variant=variant,
                        values=values,
                        provenance=Provenance(
                            sources=(
                                SourceRef(
                                    workbook=os.path.basename(self.workbook_path),
                                    sheet=sheet_name,
                                    range=(
                                        f"P{qh_row}/P{qc_row}"
                                        if "Qc" in values
                                        else f"P{qh_row}"
                                    ),
                                ),
                            ),
                            note=(
                                "Heizwärmebedarf Qh, eff (kWh/m2a) und Klimakältebedarf "
                                "Qc (kWh/m2a), Energiebilanz mit mechanischer Lüftung"
                            ),
                        ),
                    )
                )
            # Coefficient block (rows 42-63, column D); Thetai (r58) excluded
            coeff_rows = {
                "fE": (42, "-"),
                "a0": (44, "-"),
                "tau0": (45, "h"),
                "b_erdreich_boden": (55, "-"),
                "b_unbeheizt_boden": (56, "-"),
                "b_erdreich_wand": (57, "-"),
                "delta_thetah_max": (59, "°C"),
                "hue_m": (60, "m"),
                "f_horizont": (61, "-"),
                "q_vorlauf_max": (62, "°C"),
                "dq_regelung": (63, "°C"),
            }
            coeff_values: dict[str, Quantity] = {}
            for key, (row_number, unit) in coeff_rows.items():
                value = matrix[row_number - 1][3]  # D column
                if isinstance(value, (int, float)):
                    coeff_values[key] = Quantity(float(value), unit)
            if coeff_values:
                coefficients.append(
                    Sia3801Coefficients(
                        variant=variant,
                        category="",
                        coefficients=coeff_values,
                        provenance=Provenance(
                            sources=(
                                SourceRef(
                                    workbook=os.path.basename(self.workbook_path),
                                    sheet=sheet_name,
                                    range="D42:D63",
                                ),
                            ),
                            note=(
                                "SIA 380/1 Randbedingungen (fE, a0, tau0, b, "
                                "deltaThetahmax, HueM, F, q, Dq); Thetai formelabhängig"
                            ),
                        ),
                    )
                )
        return results, coefficients


def _normalize_label(label: str) -> str:
    return re.sub(r"\s+", " ", label).strip().lower()


def _normalize_station_name(name: str) -> str:
    """Normalized station-name join key (hyphens/spaces are ignored).

    The workbook spells station 5 "Bern Liebefeld" in ``Monatswerte`` but
    "Bern-Liebefeld" in ``Winter_Auslegung``; the join must tolerate that.
    """
    return re.sub(r"[\s-]+", "", name).lower()


def _column_name(index: int) -> str:
    """1-based column index -> Excel letters (4 -> "D", 120 -> "DP")."""
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _read_matrix(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> list[list[Any]]:
    """One deterministic pass over a (possibly read-only) worksheet -> value rows.

    openpyxl read-only worksheets cannot seek: every ``iter_rows`` call
    re-parses the sheet XML from the start, so the naive per-cell access
    pattern is quadratic in the number of lookups and effectively hangs on
    the wide real-workbook tables (``Qhc_Klimastat``: 483 columns).  Reading
    the requested range once into memory keeps extraction linear.  Sheets
    shorter than the requested range are padded with empty rows so callers
    can index the full range (synthetic test workbooks).
    """
    matrix = [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    ]
    expected = max_row - min_row + 1
    if len(matrix) < expected:
        width = max_col - min_col + 1
        matrix.extend([None] * width for _ in range(expected - len(matrix)))
    return matrix


def _category_codes(matrix: list[list[Any]], first_col: int, last_col: int) -> dict[int, str]:
    """Row-1 category codes of a block's columns -> ``{column: code}``."""
    codes: dict[int, str] = {}
    for col in range(first_col, last_col + 1):
        cell = matrix[0][col - 1] if col - 1 < len(matrix[0]) else None
        if cell is not None and str(cell).strip():
            codes[col] = str(cell).strip()
    return codes


def _category_block_rows(
    matrix: list[list[Any]],
    label_col: int,
    unit_col: int,
    first_cat: int,
    last_cat: int,
    stop_at_header: bool,
) -> tuple[dict[int, str], list[tuple[int, str, str, list[tuple[int, float]]]]]:
    """Metric rows 3-47 of one per-category block.

    Returns the block's category codes (row 1) and one entry per metric row:
    ``(row_number, label, unit, [(category_col, value), ...])``.  Rows without
    at least one cached numeric value are skipped (section headers, note
    rows); when ``stop_at_header`` is set the scan stops at the first
    redundant section header (e.g. "SIA 2024 Zielwerte").
    """
    categories = _category_codes(matrix, first_cat, last_cat)
    if not categories:
        return {}, []
    metric_rows: list[tuple[int, str, str, list[tuple[int, float]]]] = []
    for row_number in range(3, 48):
        row = matrix[row_number - 1] if row_number - 1 < len(matrix) else [None] * 1
        label_cell = row[label_col - 1] if label_col - 1 < len(row) else None
        if (
            stop_at_header
            and label_cell is not None
            and str(label_cell).strip()
            and _normalize_label(str(label_cell)) in _SECTION_HEADER_LABELS
        ):
            break
        values = [
            (col, row[col - 1])
            for col in categories
            if col - 1 < len(row) and isinstance(row[col - 1], (int, float))
        ]
        if not values:
            continue
        label = str(label_cell).strip() if label_cell is not None and str(label_cell).strip() else f"row{row_number}"
        unit = _category_table_unit(row[unit_col - 1] if unit_col - 1 < len(row) else None)
        metric_rows.append((row_number, label, unit, values))
    return categories, metric_rows


def _category_table_unit(raw: Any) -> str:
    """One unit cell of a category table -> a registry symbol (``"-"`` fallback).

    Unknown spellings (paren variants of the flow units, error cells) fall
    back to ``"-"`` after a small alias table; the workbook's rich-text
    superscripts are handled by :class:`Unit` itself.
    """
    if raw is None:
        return "-"
    text = str(raw).strip()
    if not text or text in _ERROR_VALUE_STRINGS or text in ("-", "–", "—", "−"):
        return "-"
    try:
        return Unit(text).symbol
    except UnitError:
        pass
    normalized = (
        text.translate(_SUPERSCRIPT_TRANSLATION)
        .replace("−", "-")
        .replace("–", "-")
        .replace("×", "x")
        .replace(" ", "")
        .replace("\u00a0", "")
    )
    return _TABLE_UNIT_ALIASES.get(normalized, "-")


def _cell(ws: Any, row: int, column: int) -> Any:
    """Read one cell from a (possibly read-only) worksheet, without cell indexing."""
    for row_values in ws.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column):
        for cell in row_values:
            return cell.value
    return None


def _safe_unit(raw: str) -> str:
    """Normalize a workbook unit cell; typographic dashes / unknown symbols -> ``"-"``."""
    unit = raw.strip()
    if unit in ("-", "–", "—", "−"):
        return "-"
    try:
        Unit(unit)
        return unit
    except UnitError:
        return "-"


def _slugify(label: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", label).strip("-").lower()
    return slug or "parameter"


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip() in ("1", "ja", "true", "TRUE", "wahr")
    return False


def _infer_data_type(values_raw: tuple[Any, Any, Any]) -> str:
    cleaned = [_clean_cell_value(raw) for raw in values_raw]
    if all(value is None for value in cleaned):
        return "text"  # no values (structural rows): nothing to type
    numeric = True
    enum_like = True
    for value in cleaned:
        if value is None:
            continue
        if not isinstance(value, (int, float)):
            numeric = False
        if isinstance(value, str) and value.lower() not in ("ja", "nein"):
            enum_like = False
    if numeric:
        return "number"
    if enum_like:
        return "enum"
    return "text"


def _clean_cell_value(value: Any) -> float | int | str | bool | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in _ERROR_VALUE_STRINGS or stripped == "":
            return None
        return stripped
    return value


def _schedule_cell(value: Any) -> float:
    """One schedule matrix cell -> fraction (missing/blank = 0.0, ``"80%"`` -> 0.8)."""
    if value is None:
        return 0.0
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped in _ERROR_VALUE_STRINGS:
            return 0.0
        if stripped.endswith("%"):
            try:
                return float(stripped[:-1]) / 100.0
            except ValueError:
                return 0.0
        try:
            return float(stripped)
        except ValueError:
            return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _month_value(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None
